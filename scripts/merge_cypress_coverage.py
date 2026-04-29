#!/usr/bin/env python3
"""
merge_cypress_coverage.py — Merge the latest auto-extracted bucket CSVs with
the manually-curated "Full Cypress coverage" files in ~/Downloads.

Purpose:
  - Carry forward the human-assigned `Assignee` and `Coverage Status`
    columns from the Downloads files.
  - Refresh `cypress_test_status`, `prod_used`, `latest_prod_timestamp`
    with the latest values from extract_features.py output.
  - For NEW rows (present in our CSV but not in Downloads), keep
    Assignee / Coverage Status BLANK — those need human assignment.
  - For STALE rows (in Downloads but no longer in our CSV — e.g. removed
    features like Preprocessing Flow), drop them.

Outputs:
  Three merged CSVs in repo root:
    - merged_b1_full_cypress_coverage.csv
    - merged_b2_full_cypress_coverage.csv
    - merged_b3_full_cypress_coverage.csv

NOTE: The dashboard does NOT read these files. It still uses the auto-
detected `cypress_test_status` from extract_features.py output. The
manually-assigned `Coverage Status` is preserved here for human review
only.
"""

import os
import csv
import sys
import sqlite3

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DB_PATH = None  # set in main()

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HOME = os.path.expanduser("~")
DOWNLOADS = os.path.join(HOME, "Downloads")

DOWNLOAD_B1 = os.path.join(DOWNLOADS, "Full Cypress coverage - connector x connector flow.csv")
DOWNLOAD_B2 = os.path.join(DOWNLOADS, "Full Cypress coverage - connector x payment method x payment method flow.csv")
DOWNLOAD_B3 = os.path.join(DOWNLOADS, "Full Cypress coverage - core flow.csv")

CURRENT_B1 = os.path.join(REPO_ROOT, "bucket_1_connector_features.csv")
CURRENT_B2 = os.path.join(REPO_ROOT, "bucket_2_connector_pm_features.csv")
CURRENT_B3 = os.path.join(REPO_ROOT, "bucket_3_core_features.csv")

OUT_B1 = os.path.join(REPO_ROOT, "merged_b1_full_cypress_coverage.csv")
OUT_B2 = os.path.join(REPO_ROOT, "merged_b2_full_cypress_coverage.csv")
OUT_B3 = os.path.join(REPO_ROOT, "merged_b3_full_cypress_coverage.csv")


def load_csv(path):
    if not os.path.exists(path):
        return [], []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return reader.fieldnames or [], list(reader)


def index_by(rows, key_fields):
    """Index rows by a tuple of values from key_fields."""
    idx = {}
    for r in rows:
        key = tuple(r.get(k, "") for k in key_fields)
        idx[key] = r
    return idx


def write_xlsx(rows, fields, out_path):
    """Write rows to xlsx with bold header row. No-op if openpyxl is missing."""
    if not HAS_OPENPYXL:
        print(f"  (skipping xlsx for {os.path.basename(out_path)} — openpyxl not installed)", file=sys.stderr)
        return
    wb = Workbook()
    ws = wb.active
    ws.append(fields)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for r in rows:
        ws.append([r.get(f, "") for f in fields])
    wb.save(out_path)


# Columns we always carry forward from the Downloads file when a row matches.
# These represent human-curated state (prod_used, prod timestamp, Assignee,
# Coverage Status) that the auto-extracted CSV doesn't have or has stale data for.
CARRY_FORWARD = ("prod_used", "latest_prod_timestamp", "Assignee", "Coverage Status")


def merge_bucket(current_path, download_path, out_path, key_fields, extra_cols):
    """
    For each row in the current (auto-extracted) CSV:
      - Look up the matching row in the download CSV by key_fields.
      - If matched: overwrite prod_used, latest_prod_timestamp, Assignee,
        Coverage Status with the Downloads values (those are the source of
        truth for the human-managed state).
      - If not matched: leave human-curated columns blank.
    Rows that are ONLY in the download file (not in current) are dropped —
    they represent features that are no longer in the codebase.
    """
    cur_fields, cur_rows = load_csv(current_path)
    dl_fields, dl_rows = load_csv(download_path)

    if not cur_rows:
        print(f"  WARN: {current_path} empty/missing", file=sys.stderr)
        return

    dl_idx = index_by(dl_rows, key_fields)

    # Output schema = current schema + every CARRY_FORWARD column not already
    # there + any extra columns from the Downloads file (Assignee, Coverage
    # Status). The current auto-extracted CSV doesn't have prod_used or
    # latest_prod_timestamp; we explicitly add them so values from Downloads
    # can flow into the merged output.
    extras_in_output = []
    for c in list(CARRY_FORWARD) + list(extra_cols):
        if c not in cur_fields and c not in extras_in_output:
            extras_in_output.append(c)
    output_fields = list(cur_fields) + extras_in_output

    matched = 0
    new_rows = 0
    pulled_cols = list(CARRY_FORWARD)
    merged_rows = []

    for r in cur_rows:
        key = tuple(r.get(k, "") for k in key_fields)
        merged = dict(r)
        dl_match = dl_idx.get(key)
        if dl_match:
            matched += 1
            # Pull every carry-forward column we know about from Downloads,
            # overwriting whatever the auto-extract had.
            for c in pulled_cols:
                if c in dl_match:
                    merged[c] = dl_match[c]
            for c in extra_cols:
                if c not in merged:
                    merged[c] = dl_match.get(c, "")
        else:
            new_rows += 1
            for c in extra_cols:
                merged.setdefault(c, "")
        # Make sure every output column is present (even if empty)
        for c in output_fields:
            merged.setdefault(c, "")
        merged_rows.append(merged)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=output_fields)
        writer.writeheader()
        writer.writerows(merged_rows)

    xlsx_path = out_path.replace(".csv", ".xlsx")
    write_xlsx(merged_rows, output_fields, xlsx_path)

    stale = len(dl_rows) - matched
    print(f"  {os.path.basename(out_path)} (+ .xlsx): {len(cur_rows)} rows total "
          f"({matched} matched from Downloads → pulled prod/Assignee/Coverage Status, "
          f"{new_rows} new with blanks, {stale} stale dropped)")


def regenerate_report():
    """
    Rewrite feature_extraction_report.{csv,xlsx} with the manually-curated
    columns folded in alongside the auto-detected stats.

    Adds columns:
      - assigned        : rows where the merged file has a non-empty Assignee
      - manual_covered  : rows where Coverage Status == "Covered"
      - manual_review   : rows where Coverage Status == "Review"
      - manual_blocked  : rows where Coverage Status indicates blocked work
                         (Creds not available, Sandbox creds is not there,
                          Not possible, UCS only connector,
                          Internal subflow should be ignored)
    """
    bucket_files = [
        ("Bucket 1", OUT_B1, "Connector × Feature (payload-level, PM-agnostic)"),
        ("Bucket 2", OUT_B2, "Connector × PM × PMT × Feature"),
        ("Bucket 3", OUT_B3, "Core features (connector-agnostic)"),
    ]

    blocked_statuses = {
        "Creds not available",
        "Sandbox creds is not there",
        "Not possible",
        "UCS only connector",
        "Internal subflow should be ignored",
    }

    rows_out = []
    totals = {
        "total": 0, "cy_covered": 0, "cy_not_covered": 0, "cy_no_config": 0,
        "prod_yes": 0, "prod_no": 0, "prod_unknown": 0,
        "assigned": 0, "manual_covered": 0, "manual_review": 0, "manual_blocked": 0,
    }
    for bucket, fname, desc in bucket_files:
        if not os.path.exists(fname):
            print(f"  WARN: {fname} not found — run merge first", file=sys.stderr)
            continue
        cy = {"covered": 0, "not_covered": 0, "no_cypress_config": 0}
        pr = {"yes": 0, "no": 0, "unknown": 0}
        assigned = manual_covered = manual_review = manual_blocked = 0
        with open(fname) as f:
            rows = list(csv.DictReader(f))
        for r in rows:
            cy[r.get("cypress_test_status", "")] = cy.get(r.get("cypress_test_status", ""), 0) + 1
            pr[r.get("prod_used", "unknown") or "unknown"] = pr.get(r.get("prod_used", "unknown") or "unknown", 0) + 1
            if (r.get("Assignee") or "").strip():
                assigned += 1
            cs = (r.get("Coverage Status") or "").strip()
            if cs == "Covered":
                manual_covered += 1
            elif cs == "Review":
                manual_review += 1
            elif cs in blocked_statuses:
                manual_blocked += 1
        total = len(rows)
        cy_pct = f"{round(cy['covered'] / total * 100)}%" if total else "0%"
        denom = pr["yes"] + pr["no"]
        pr_pct = f"{round(pr['yes'] / denom * 100)}%" if denom else "0%"

        rows_out.append([
            bucket, desc, total,
            cy["covered"], cy["not_covered"], cy["no_cypress_config"], cy_pct,
            pr["yes"], pr["no"], pr["unknown"], pr_pct,
            assigned, manual_covered, manual_review, manual_blocked,
        ])
        totals["total"] += total
        totals["cy_covered"] += cy["covered"]
        totals["cy_not_covered"] += cy["not_covered"]
        totals["cy_no_config"] += cy["no_cypress_config"]
        totals["prod_yes"] += pr["yes"]; totals["prod_no"] += pr["no"]; totals["prod_unknown"] += pr["unknown"]
        totals["assigned"] += assigned
        totals["manual_covered"] += manual_covered
        totals["manual_review"] += manual_review
        totals["manual_blocked"] += manual_blocked

    if not rows_out:
        return

    cy_pct_total = f"{round(totals['cy_covered'] / totals['total'] * 100)}%" if totals["total"] else "0%"
    denom_t = totals["prod_yes"] + totals["prod_no"]
    pr_pct_total = f"{round(totals['prod_yes'] / denom_t * 100)}%" if denom_t else "0%"
    rows_out.append([
        "TOTAL", "All buckets combined", totals["total"],
        totals["cy_covered"], totals["cy_not_covered"], totals["cy_no_config"], cy_pct_total,
        totals["prod_yes"], totals["prod_no"], totals["prod_unknown"], pr_pct_total,
        totals["assigned"], totals["manual_covered"], totals["manual_review"], totals["manual_blocked"],
    ])

    headers = [
        "bucket", "description", "total_rows",
        "cypress_covered", "cypress_not_covered", "cypress_no_config", "cypress_coverage_pct",
        "prod_used_yes", "prod_used_no", "prod_used_unknown", "prod_usage_pct",
        "assigned", "manual_covered", "manual_review", "manual_blocked",
    ]

    csv_path = os.path.join(REPO_ROOT, "feature_extraction_report.csv")
    with open(csv_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows_out)

    if HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
        for r in rows_out:
            ws.append(r)
        wb.save(csv_path.replace(".csv", ".xlsx"))

    print(f"  feature_extraction_report.{{csv,xlsx}}: {len(rows_out)} bucket rows "
          f"(now includes assigned / manual_covered / manual_review / manual_blocked)")


def sync_to_db(merged_files):
    """
    Push the human-curated columns (assignee, coverage_status) from the
    merged CSVs into the issues table — but ONLY for cells where the DB
    is currently NULL/empty. This prevents overwriting edits made through
    the web app.
    """
    db = os.path.join(REPO_ROOT, "features.db")
    if not os.path.exists(db):
        print("  (skipping DB sync — features.db not found)")
        return
    conn = sqlite3.connect(db)
    if not [r for r in conn.execute("PRAGMA table_info(issues)").fetchall() if r[1] == "assignee"]:
        print("  (skipping DB sync — old issues schema; run extract_features.py first)")
        conn.close()
        return

    total_synced = 0
    for path, bucket, key_fields in merged_files:
        if not os.path.exists(path):
            continue
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        for r in rows:
            assignee = (r.get("Assignee") or "").strip() or None
            coverage_status = (r.get("Coverage Status") or "").strip() or None
            if assignee is None and coverage_status is None:
                continue
            key_vals = {k: (r.get(k) or "") for k in key_fields}
            conditions = ["bucket = ?"]
            params = [bucket]
            for col in ("connector", "pm", "pmt"):
                conditions.append(f"{col} = ?")
                params.append(key_vals.get(col, ""))
            conditions.append("feature = ?")
            params.append(key_vals.get("feature", ""))

            cur = conn.execute(
                f"""UPDATE issues
                    SET assignee        = COALESCE(NULLIF(assignee, ''),        ?),
                        coverage_status = COALESCE(NULLIF(coverage_status, ''), ?),
                        updated_at      = CURRENT_TIMESTAMP
                    WHERE {' AND '.join(conditions)}
                      AND (assignee IS NULL OR assignee = ''
                           OR coverage_status IS NULL OR coverage_status = '')""",
                [assignee, coverage_status] + params,
            )
            total_synced += cur.rowcount
    conn.commit()
    conn.close()
    print(f"  features.db: {total_synced} rows received initial assignee/coverage_status from Downloads")


def main():
    print(f"Loading current CSVs from {REPO_ROOT}", file=sys.stderr)
    print(f"Loading Downloads files from {DOWNLOADS}", file=sys.stderr)
    print()

    merge_bucket(
        CURRENT_B1, DOWNLOAD_B1, OUT_B1,
        key_fields=("connector", "feature"),
        extra_cols=("Assignee", "Coverage Status"),
    )

    merge_bucket(
        CURRENT_B2, DOWNLOAD_B2, OUT_B2,
        key_fields=("connector", "payment_method", "payment_method_type", "feature"),
        extra_cols=("Assignee",),
    )

    merge_bucket(
        CURRENT_B3, DOWNLOAD_B3, OUT_B3,
        key_fields=("feature",),
        extra_cols=("Assignee", "Coverage Status"),
    )

    regenerate_report()

    sync_to_db([
        (OUT_B1, 1, ("connector", "feature")),
        (OUT_B2, 2, ("connector", "payment_method", "payment_method_type", "feature")),
        (OUT_B3, 3, ("feature",)),
    ])

    print()
    print("Done. Merged outputs:")
    for p in (OUT_B1, OUT_B2, OUT_B3):
        print(f"  {p} (+ .xlsx)")
    print(f"  {os.path.join(REPO_ROOT, 'feature_extraction_report.csv')} (+ .xlsx)")
    print()
    print("NOTE: The dashboard still uses the auto-detected cypress_test_status,")
    print("not the manually-curated 'Coverage Status' column. These merged files")
    print("are for human review/triage only.")


if __name__ == "__main__":
    main()
